import openpyxl
path="C:\\Users\\hotpu\\OneDrive\\Desktop\\bank.xlsx"
wb=openpyxl.load_workbook(path)
sheet=wb.active
user_name=input("Enter the username")
user_pwd=input("Enter the password")
i=2
while True:
    if sheet.cell(row=i,column=1).value==user_name and sheet.cell(row=i,column=2).value==user_pwd:
        print("Choice 1 to check balance")
        print("Choice 2 for withdrawal")
        print("Choice 3 for deposit")
        print("Choice 4 for transfer of money from one number to other")
        choice=int(input("Enter your choice"))
        if choice==1:
            s=sheet.cell(row=i, column=3).value
            print(s)

        elif choice==2:
            wit=int(input("Enter the amount to be withdrawn"))
            if wit>sheet.cell(row=i,column=3).value:
                print("Insufficient balance")
            else:
                a=sheet.cell(row=i,column=3).value
                a-=wit
                sheet.cell(row=i, column=3).value=a
                wb.save(path)


        elif choice==3:
            dep = int(input("Enter the amount to be deposited"))
            if dep> sheet.cell(row=i, column=3).value:
                print("Insufficient balance")
            b=sheet.cell(row=i, column=3).value
            b+=dep
            sheet.cell(row=i, column=3).value=b
            wb.save(path)

        elif choice==4:
           user_no=input("Enter the mobile number to which the amount is to be transfered")
           transfer=input("Enter the amount")
           if sheet.cell(row=i, column=4).value == user_no:
               d=sheet.cell(row=i, column=3).value
               d-=transfer
               sheet.cell(row=i, column=3).value=d



        break
    elif sheet.cell(row=i,column=1).value==None and sheet.cell(row=i,column=2).value==None:
        print("Invalid")
        break
    i+=1